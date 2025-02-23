import csv
import re
from pathlib import Path
from collections import defaultdict

import openpyxl
import pandas as pd
import numpy as np
import math
import random
from inflection import singularize

from django.template.defaultfilters import slugify
from django.contrib.contenttypes.models import ContentType
from django.utils import timezone

from sequences import get_next_value

from .models import Domain, Section, Requirement, Constraint, Category, Statement, ConstraintStatement, DataManagementPlan, ConstraintDependency, Definition
from .entity_types import EntityType
from workflows.tenant_models import Tenant
from mir.models import DataManagement

def strip(s):
    return re.sub(r"^[\s\n\t]*(.*)[\s\n\t\r]*", r"\1", s)

def isnan(s):
    if s == None:
        return True
    if s == "nan":
        return True
    if s == np.nan:
        return True
    if isinstance(s, float) and math.isnan(s):
        return True
    if isinstance(s, str) and s == "":
        return True

def get_domain(tenant_id, name, description):
    slug = slugify(name)
    domain, _ = Domain.unscoped.update_or_create(tenant_id=tenant_id, name=name, defaults={"slug": slug, "description": description})
    return domain

def create_data_managements(tenant, domain):
        for i, name in enumerate("Routine Task".split()):
            content_type = ContentType.objects.get(app_label='workflows', model=name.lower())
            dm = DataManagement.objects.create(tenant_id=tenant.id, index=i, content_type=content_type, policy=DataManagement.POLICY_MANAGED, allow_policy_change=False, status=DataManagement.STATUS_PROTOTYPING)
            DataManagementPlan.objects.create(tenant_id=tenant.id, domain_id=domain.id, index=i, plan=DataManagementPlan.PLAN_IMPLEMENT, data_management=dm)

        for i, name in enumerate("Employee Training TrainingOrganized TrainingAttended".split(), start=10):
            content_type = ContentType.objects.get(app_label='mir', model=name.lower())
            dm = DataManagement.objects.create(tenant_id=tenant.id, index=i, content_type=content_type, policy=DataManagement.POLICY_MANAGED, status=DataManagement.STATUS_PROTOTYPING)
            DataManagementPlan.objects.create(tenant_id=tenant.id, domain_id=domain.id, index=i, plan=DataManagementPlan.PLAN_IMPLEMENT, data_management=dm)

colors = [
    "#0d6efd", # blue
    "#6610f2", # indigo
    "#6f42c1", # purple
    "#d63384", # pink
    "#dc3545", # $red
    "#fd7e14", # $orange
    "#ffc107", # $yellow
    "#198754", # $green
    "#20c997", # $teal
    "#0dcaf0", # $cyan
    "#cccccc", # $gray
]

def get_or_create_category(tenant, domain, name, parent=None, index=None):
    try:
        cat = Category.unscoped.get(tenant_id=tenant.id, domain_id=domain.id, name=name)
        return cat
    except:
        pass

    if not index:
        index = get_next_value('categories')
    if parent:
        color = parent.color
    else:
        color = colors[index % len(colors)]
    return Category.unscoped.create(tenant_id=tenant.id, domain_id=domain.id, name=name, index=index, color=color, parent=parent)
    
def create_categories(tenant, domain):
    categories = {}
    for i, name in enumerate("DOC CTM CHM MIR INC OPS RIM TECH SDC ORG SERV".split()):
        categories[name] = get_or_create_category(tenant, domain, name)

    get_or_create_category(tenant, domain, "No category", index=32767)

    return categories

def get_docid_path(docid):
    """
    Return all docids
    """
    p = []
    for d in docid.split('.'):
        if len(p) > 0:
            dd = p[-1] + '.' + d
        else:
            dd = d
        p.append(dd)

    return p

def get_entity_type(term):
    a = term.split()
    if len(a) > 0:
        return EntityType.get_by_name(a[-1])
    return None
        
def create_definitions(tenant, domain, constraint):
    for term in re.findall(r'\$([^$]*)\$', constraint.text):
        entity_type = get_entity_type(term) or EntityType.NOT_DEFINED
        try:
            definition = Definition.unscoped.get(tenant_id=tenant.id, domain_id=domain.id, term__iexact=term)
        except Definition.DoesNotExist:
            index = get_next_value('definitions')
            definition = Definition.unscoped.create(tenant_id=tenant.id, domain_id=domain.id, term=term.capitalize(), index=index, ref_entity_type=entity_type)

        constraint.definitions.add(definition)

    for term_plural in re.findall(r'\@([^@]*)\@', constraint.text):

        term = singularize(term_plural)
        entity_type = get_entity_type(term) or EntityType.NOT_DEFINED
        try:
            definition = Definition.unscoped.get(tenant_id=tenant.id, domain_id=domain.id, term__iexact=term)
        except Definition.DoesNotExist:
            index = get_next_value('definitions')
            definition = Definition.unscoped.create(tenant_id=tenant.id, domain_id=domain.id, term=term.capitalize(), index=index, ref_entity_type=entity_type, ref_plural=True, term_plural=term_plural)

        constraint.definitions.add(definition)

def create_dependencies(tenant, domain, constraint, deps):
    if not isnan(deps):
        deps = deps
        for cid in [ x.strip() for x in deps.strip().split(',')]:
            dep = Constraint.unscoped.filter(tenant_id=tenant.id, domain_id=domain.id, key=cid).first()
            constraint.dependencies.add(dep)

def import_excel(tenant, path):
    now = timezone.now()

    workbook = openpyxl.load_workbook(path)

    ws = workbook.get_sheet_by_name('USM.TOOLS')

    for row in ws.iter_rows(min_row=1, max_row=1):
        d = row[0].value

    for row in ws.iter_rows(min_row=2, max_row=2):
        headers = [cell.value for cell in row]

    if m := re.search(r'USM.TOOLS#NAME:"((?:(?!(?<!\\)").)*)",DESCRIPTION:"((?:(?!(?<!\\)").)*)', d):
        domain = get_domain(tenant.id, name=m.group(1), description=m.group(2))
    else:
        raise ValueError(_("This is not Excel file supported by usm.tools"))

    categories = create_categories(tenant, domain)
    create_data_managements(tenant, domain)

    parent_section = None

    section = None
    requirement = None
    statement = None
    constraint = None
    docids = None
    p = {}

    for row in ws.iter_rows(min_row=3, values_only=True):
        e = dict(zip(headers, row))

        if isnan(e["C. ID"]):
            break

        for k, v in e.items():
            if v == "-":
                e[k] = None
            elif not v and k in p:
                e[k] = p[k]
        p = e

        if e['Doc ID']:
            if e['Doc ID'] == '-':
                docids = []
            else:
                docids = get_docid_path(e['Doc ID'])

        for i, docid in enumerate(docids):

            header_name = (('sub-' * docid.count('.')) + 'section title').capitalize()

            title = e[header_name]

            if i == len(docids) - 1:
                text = e["Section text"]
            else:
                text = None

            if title == "-":
                break
                section = None

            index = docid.split('.')[-1]

            section, created = Section.unscoped.get_or_create(
                                    tenant=tenant, 
                                    domain_id=domain.id,
                                    doc=e["Doc"],
                                    docid=docid,
                                    defaults={"title": title, "parent": parent_section, "text": text, "index": index})

            parent_section = section

        if e["Requirement"] == "-":
            requirement = None
        else:
            requirement, _ = Requirement.unscoped.get_or_create(tenant_id=tenant.id, section_id=section.id, text=e["Requirement"], defaults={"index": get_next_value('requirement')})


        if not isnan(e["S. Title"]):
            if e["S. Title"] == "-":
                statement = None
            else:
                statement, _ = Statement.unscoped.get_or_create(tenant_id=tenant.id, requirement_id=requirement.id, title=e["S. Title"], defaults={"text": e["S. Text"], 'index': get_next_value('statement')})
        else:
            assert statement, "statement must be defined"

        if e["C. ID"] == "-":
            constraint = None
        else:
            is_generic = e["C. Generic"] == "Yes"
            top_category = categories[e["C. Deployment category"]]
            if not isnan(e["C. Sub-category"]):
                category = get_or_create_category(tenant, domain, e["C. Sub-category"], parent=top_category)
            else:
                category = top_category
            constraint, created = Constraint.unscoped.get_or_create(tenant=tenant, domain_id=domain.id, key=e["C. ID"].strip(), defaults={"title": e["C. Title"], "text": e["C. Text"], "story_points": float(e["C. Story points"]), "is_generic": is_generic, "category": category, "index": get_next_value('constraint')})

            if created:
                ConstraintStatement(constraint=constraint, statement=statement).save()

            create_dependencies(tenant, domain, constraint, e["C. Dependencies"])
            create_definitions(tenant, domain, constraint)

        p = e.copy()
